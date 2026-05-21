"""
CRM Service - Business Logic Layer

This file contains the "business logic" for CRM operations.
Think of it as the "middle layer" between:
- API routes (what users call)
- Database (where data is stored)

Why separate? 
- Keeps API routes clean and simple
- Makes business logic reusable
- Easier to test
"""
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from uuid import UUID as UUIDType, uuid5, NAMESPACE_URL
import json
import re
import tempfile
import logging
from pathlib import Path

from supabase import Client
from thefuzz import fuzz
from openpyxl import load_workbook

from app.database.connection import get_supabase_client, get_supabase_service_client
from app.models.crm import (
    Customer,
    CustomerCreate,
    CustomerUpdate,
    Interaction,
    InteractionCreate,
    InteractionUpdate,
    QuoteDraftRequest,
    DashboardMetrics,
    QuietCustomerSummary,
    WeeklyInteractionCount,
    CustomerProfileUpdate,
    CustomerProfileFeedbackCreate,
    CustomerProfileFeedback,
)
from app.services.ai_service import (
    gemini_chat,
    gemini_embed,
    log_conversation_to_rag,
)
from app.services.conversation_archive_service import (
    _parse_chatgpt_export_row,
    get_chatgpt_export_archives_for_customer,
)
from app.services.profile_research_service import (
    PROFILE_CONTEXT_MAX_CHARS,
    build_profile_research_context,
    gather_profile_research_inputs,
)
from app.services.telegram_service import notify_interaction_saved

# Sales stage definitions (Brian Tracy 7-stage process)
SALES_STAGES = {
    "1": "Prospecting",
    "2": "Rapport",
    "3": "Needs Analysis",
    "4": "Presenting Solution",
    "5": "Handling Objections",
    "6": "Closing",
    "7": "Follow-up & Cross-sell",
}
from app.services.pms_service import get_all_categories
from app.utils.profile_text import sanitize_profile_plain_text


# =============================
# CUSTOMER SERVICES
# =============================


def get_all_customers(
    limit: int = 100,
    offset: int = 0,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> List[Customer]:
    """Get all customers from the database with pagination.
    
    If date filters are provided, only returns customers that have interactions
    within the specified date range.
    
    Args:
        limit: Maximum number of customers to return
        offset: Number of customers to skip
        start_date: Optional ISO date string (YYYY-MM-DD) - filter customers with interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter customers with interactions up to this date
    """
    supabase: Client = get_supabase_client()

    # If date filters are provided, get customers that have interactions in that range
    if start_date or end_date:
        # First, get distinct customer_ids from interactions table with date filter
        interaction_query = supabase.table("interactions").select("customer_id")
        
        if start_date:
            interaction_query = interaction_query.gte("created_at", f"{start_date}T00:00:00")
        if end_date:
            interaction_query = interaction_query.lte("created_at", f"{end_date}T23:59:59")
        
        interactions_res = interaction_query.execute()
        customer_ids = list(set(row.get("customer_id") for row in (interactions_res.data or [])))
        
        if not customer_ids:
            return []
        
        # Then fetch those customers
        response = (
            supabase.table("customers")
            .select("*")
            .in_("customer_id", customer_ids)
            .order("created_at", desc=True)
            .limit(limit)
            .offset(offset)
            .execute()
        )
    else:
        # Normal query without date filtering
        response = (
            supabase.table("customers")
            .select("*")
            .order("created_at", desc=True)
            .limit(limit)
            .offset(offset)
            .execute()
    )

    return [Customer(**row) for row in (response.data or [])]


def get_customer_by_id(customer_id: str) -> Optional[Customer]:
    """Get a single customer by UUID.

    Also ensures that `latest_profile_text` is populated from the most recent
    AI-generated profile interaction if the column is still NULL in the DB.
    """
    supabase: Client = get_supabase_client()

    response = (
        supabase.table("customers")
        .select("*")
        .eq("customer_id", customer_id)
        .execute()
    )

    if not response.data:
        return None

    row = response.data[0]

    # If latest_profile_text is missing/empty, try to derive it from interactions
    latest_profile_text = row.get("latest_profile_text") or None

    if not latest_profile_text:
        try:
            interactions = get_interactions_for_customer(customer_id, limit=1, offset=0)
        except Exception as e:
            logging.warning(f"Failed to derive latest_profile_text from interactions for {customer_id}: {str(e)}")
            interactions = []

        if interactions:
            inter = interactions[0]
            # Prefer AI response (which usually contains the ICP), fall back to input_text
            derived_text = (inter.ai_response or inter.input_text or "").strip()
            if derived_text:
                row["latest_profile_text"] = derived_text
                # Use interaction created_at as a reasonable profile-updated timestamp if missing
                if not row.get("latest_profile_updated_at"):
                    row["latest_profile_updated_at"] = getattr(inter, "created_at", None)

    return Customer(**row)


def get_customers_count() -> int:
    """Get total number of customers in the database."""
    supabase: Client = get_supabase_client()

    response = supabase.table("customers").select("customer_id", count="exact").execute()

    return response.count if getattr(response, "count", None) is not None else 0


def search_customers_by_name(query: str, limit: int = 20) -> List[Customer]:
    """Search customers by partial name (case-insensitive).

    This uses Postgres ILIKE under the hood to match anywhere in the name.
    Example: query='sika' will match 'Sika Abyssinia', 'Sika Ethiopia', etc.
    """
    supabase: Client = get_supabase_client()

    response = (
        supabase.table("customers")
        .select("*")
        .ilike("customer_name", f"%{query}%")
        .order("customer_name", desc=False)
        .limit(limit)
        .execute()
    )

    return [Customer(**row) for row in (response.data or [])]


def _generate_display_id() -> str:
    """Generate a human-readable customer ID like LC-YYYY-CUST-0001.

    This mirrors the logic from the original Streamlit app so that
    new customers in the FastAPI backend keep a similar ID format.
    """
    supabase: Client = get_supabase_client()
    year = datetime.now().year

    # Fetch existing display IDs to find the highest counter for this year
    response = supabase.table("customers").select("display_id").execute()

    max_num = 0
    for row in response.data or []:
        display_id = row.get("display_id") or ""
        prefix = f"LC-{year}-CUST-"
        if isinstance(display_id, str) and display_id.startswith(prefix):
            try:
                num = int(display_id.split("-")[-1])
                if num > max_num:
                    max_num = num
            except ValueError:
                continue

    new_num = max_num + 1
    return f"LC-{year}-CUST-{new_num:04d}"


def create_customer(customer_in: CustomerCreate) -> Customer:
    """Create a new customer.

    If display_id is not provided, a new one is generated automatically.
    """
    supabase: Client = get_supabase_client()

    # ---------------------------------------------
    # 1) Duplicate-check inspired by Streamlit v6
    # ---------------------------------------------
    # Look for customers with similar names using a fuzzy match.
    # In the original Streamlit app this was a separate step in the UI;
    # here we surface it as a clear error so the frontend can warn the user.
    existing_response = (
        supabase.table("customers")
        .select("customer_id, customer_name, display_id")
        .ilike("customer_name", f"%{customer_in.customer_name}%")
        .limit(20)
        .execute()
    )
    similar_names: List[str] = []
    for row in existing_response.data or []:
        name = (row.get("customer_name") or "").strip()
        if not name:
            continue
        score = fuzz.partial_ratio(
            customer_in.customer_name.lower(),
            name.lower(),
        )
        if score >= 85:
            display_id = row.get("display_id") or "—"
            similar_names.append(f"{name} (ID: {display_id}, score: {score})")

    if similar_names:
        # Let the API layer translate this into a 409 Conflict.
        joined = "; ".join(similar_names[:3])
        raise ValueError(
            f"Similar customers already exist. Please review before creating a new one: {joined}"
        )

    # ---------------------------------------------
    # 2) Create the base customer row
    # ---------------------------------------------
    data = customer_in.model_dump()
    if not data.get("display_id"):
        data["display_id"] = _generate_display_id()

    response = supabase.table("customers").insert(data).execute()

    if not response.data:
        raise RuntimeError("Failed to create customer")

    customer = Customer(**response.data[0])
    return customer


def update_customer(customer_id: str, customer_update: CustomerUpdate) -> Customer:
    """Update an existing customer."""
    
    supabase: Client = get_supabase_client()
    
    # Check if customer exists
    existing = get_customer_by_id(customer_id)
    if not existing:
        raise ValueError("Customer not found")
    
    # Build update payload (only include fields that are provided)
    update_data = customer_update.model_dump(exclude_unset=True)
    
    if not update_data:
        # No fields to update
        return existing
    
    # If updating customer_name, check for duplicates
    if "customer_name" in update_data:
        existing_response = (
            supabase.table("customers")
            .select("customer_id, customer_name, display_id")
            .ilike("customer_name", f"%{update_data['customer_name']}%")
            .limit(20)
            .execute()
        )
        for row in existing_response.data or []:
            if str(row.get("customer_id")) != customer_id:
                name = (row.get("customer_name") or "").strip()
                if name:
                    score = fuzz.partial_ratio(
                        update_data["customer_name"].lower(),
                        name.lower(),
                    )
                    if score >= 85:
                        raise ValueError(f"Similar customer already exists: {name}")
    
    # Update the customer
    response = (
        supabase.table("customers")
        .update(update_data)
        .eq("customer_id", customer_id)
        .execute()
    )
    
    if not response.data:
        raise RuntimeError("Failed to update customer")

    return Customer(**response.data[0])


def delete_customer(customer_id: str) -> None:
    """Delete a customer and all associated interactions (cascade)."""
    supabase: Client = get_supabase_client()
    
    # Check if customer exists
    existing = get_customer_by_id(customer_id)
    if not existing:
        raise ValueError("Customer not found")
    
    # Delete customer (interactions will be cascade deleted by database foreign key)
    supabase.table("customers").delete().eq("customer_id", customer_id).execute()


def build_customer_profile(customer_id: str, user_id: Optional[str] = None) -> Customer:
    """
    Generate a customer profile using AI, existing conversations, and web search.
    
    This function replicates the exact logic from the Streamlit MVP (v6-gemini-final.py):
    - Searches relevant documents and memories (RAG)
    - Searches web for company information (Google PSE, SerpAPI, Wikipedia)
    - Searches LinkedIn for decision-makers
    - Generates comprehensive profile with Strategic-Fit Matrix
    - Uses dynamic product categories from chemical_types table
    
    This is called on-demand when the user clicks "Build Profile" button.
    """
    supabase: Client = get_supabase_client()
    
    # Ensure customer exists
    customer = get_customer_by_id(customer_id)
    if not customer:
        raise RuntimeError("Customer not found")
    
    # Unified CRM history: interactions table + conversation archive
    try:
        interactions, table_total, archive_added, pipeline_added, chatgpt_added = (
            merge_customer_interaction_history(str(customer.customer_id))
        )
        logging.info(
            "Profile build for %s: %s merged history (%s interactions + %s conversation + %s pipeline + %s ChatGPT export)",
            customer.customer_name,
            len(interactions),
            table_total,
            archive_added,
            pipeline_added,
            chatgpt_added,
        )
    except Exception as e:
        logging.warning(
            f"Failed to fetch interactions for customer {customer.customer_id}: {str(e)}"
        )
        interactions = []
        table_total = 0
        archive_added = 0
        pipeline_added = 0
        chatgpt_added = 0

    research_inputs = gather_profile_research_inputs(
        customer, interactions, user_id=user_id
    )
    context, research_meta = build_profile_research_context(
        customer,
        rag_docs=research_inputs["rag_docs"],
        interactions=research_inputs["interactions"],
        web_context=research_inputs["web_context"],
        linkedin_context=research_inputs["linkedin_context"],
        conversation_logs=None,
    )
    logging.info(
        "Profile context for %s: %s chars sent (raw %s, max %s) | RAG=%s CRM=%s web=%s linkedin=%s",
        customer.customer_name,
        research_meta.get("total_research_chars_sent"),
        research_meta.get("total_research_chars_raw"),
        PROFILE_CONTEXT_MAX_CHARS,
        research_meta.get("rag_document_count"),
        research_meta.get("crm_interaction_count"),
        research_meta.get("web_search_chars"),
        research_meta.get("linkedin_chars"),
    )

    # Step 5: Fetch unique categories from chemical_types table (dynamically)
    try:
        categories_list = get_all_categories()
        if not categories_list:
            categories_list = ["Cement", "Dry-Mix", "Admixtures", "Paint & Coatings"]
    except Exception as e:
        logging.warning(f"Failed to fetch categories: {str(e)}. Using defaults.")
        categories_list = ["Cement", "Dry-Mix", "Admixtures", "Paint & Coatings"]
    
    # Build category list for prompt (normalize keys for JSON: lowercase, underscores)
    category_prompt_lines = []
    category_json_keys = {}
    for cat in categories_list:
        json_key = cat.lower().replace(" ", "_").replace("-", "_").replace("&", "and")
        category_json_keys[cat] = json_key
        category_prompt_lines.append(f"- {cat} (0=No Fit, 1=Low Fit, 2=Moderate Fit, 3=High Fit)")
    
    categories_text = "\n".join(category_prompt_lines)
    json_example_keys = {json_key: "0-3" for json_key in category_json_keys.values()}
    json_example = json.dumps({"strategic_fit_matrix": json_example_keys}, indent=2)
    
    # Step 6: Create the enhanced system prompt (based on Streamlit MVP, simplified for readability)
    # NOTE: We explicitly control style so the output is clean and easy to read inside the CRM UI.
    system_prompt = f"""You are an Industry-Intel Research Assistant and B2B Chemical-Supply Strategist for LeanChem.

Write a thorough Ideal Customer Profile for the target company and every construction-relevant subsidiary in Ethiopia. The reader should follow a clear story: what we know from research, who the company is, where they manufacture in Ethiopia, how LeanChem fits, then what to do next.

LENGTH: Aim for 2,500–3,500 words when the research context is rich. Never compress CRM, RAG, web, or LinkedIn facts into a short summary—include specifics (dates, products, sites, people, quotes, objections).

If the company is a conglomerate, cover all major units relevant to construction, chemicals, and manufacturing.

LeanChem offerings (for fit reasoning):
- Dry-Mix/Plaster: RDP, HPMC, Starch Ether, Fiber, Zinc Stearate, Plasticizer, Defoamer, SBR, Acrylic Waterproofing, White Cement, Iron Oxide, Titanium Dioxide
- Concrete Admixtures: PCE, SNF, Lignosulphonate, Sodium Gluconate, Penetrol-type waterproofing
- Paint/Coatings: Styrene-Acrylic Binders, Pure Acrylics, VAE, HEC, White Cement, Iron Oxide, Titanium Dioxide
- Cement Grinding: cement grinding aids

Strategic-fit categories (score 0–3 each):
{categories_text}
0 = No Fit, 1 = Low Fit, 2 = Moderate Fit, 3 = High Fit
Base scores on volume vs LeanChem capacity, pain points LeanChem can solve (forex, lead time, performance), and switching likelihood.

MANDATORY — USE ALL RESEARCH SECTIONS IN THE USER MESSAGE:
Labeled blocks: CUSTOMER RECORD, RAG DOCUMENTS, CRM INTERACTIONS, WEB SEARCH, LINKEDIN.
Read every non-empty block. Reflect those facts in section 0 and weave them into sections 1–4. Do not invent facts missing from context.

OUTPUT FORMAT (CRITICAL):
- Plain text only: no markdown tables (no | pipes), no ### headers, no **bold**, no ``` fences, no emojis, no [text](url) links.
- Use exactly these five numbered section headings in this order:

0. Research Context Summary
Digest of all inputs before analysis. Use these subsection titles on their own lines (no # symbols), each followed by bullet lines (- item):
RAG documents — at least 5 bullets when RAG data exists; otherwise one bullet "No RAG matches".
CRM interactions — use the COMPLETE interaction index in the CRM INTERACTIONS block (every log is listed); summarize and cite specific dates, quotes, and objections from those logs. Do not skip older interactions.
Web search — company facts, sites, news, investments from the WEB SEARCH block.
LinkedIn — people found: Name, Position, full LinkedIn URL per bullet.
End with: Total research context: [N] RAG docs, [N] CRM interactions, web=[yes/no], LinkedIn=[yes/no].

1. Company Snapshot
Multi-paragraph overview: core business, group structure, scale in Ethiopia, recent news and investments (cement, dry-mix, admixtures, coatings). Name sources in prose (no [1] citation markers).

2. Construction Footprint in Ethiopia
List every relevant plant, unit, or subsidiary. One bullet per line:
- Business Unit – Construction Products – Location (City, Country) – Scale or capacity hint – Source
Cover cement, dry-mix, admixtures, and coatings where present. Say explicitly if a vertical is absent.

3. Strategic Fit Assessment
Start with 2–4 paragraphs explaining overall LeanChem opportunity and risks.
Then subsection title: Strategic-Fit Matrix
Then one line per category: CategoryName: X/3 - detailed reason (use exact category names: {', '.join(categories_list)}).
Add subsection: Score Rationale
Explain volume, pain points, competition, and switching per high-scoring category.

4. Recommended Next Steps
Subsection: Interaction Review
Summarize CRM relationship stage, blockers, and momentum from section 0 CRM bullets.
Subsection: Strategic Actions
One line per product category ({', '.join(categories_list)}), exactly this format:
CategoryName: [2–3 sentence key analysis] — [concrete next step: outreach, samples, contract, technical advisory]
Use "No direct engagement" as the next step when fit is 0/3 or there is no viable opportunity. Do not write this subsection as a single paragraph.
Subsection: Key Contacts
Up to 10 verified decision-makers, one block per person:
Name: ...
Position: ...
LinkedIn: full URL
Source: ...

CRITICAL: At the END of your response, include a JSON block with Strategic-Fit Matrix scores:
{json_example}

Use the exact category names as keys (lowercase, underscores for spaces)."""
    
    # Step 7: Build messages with context
    messages = [
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": (
                f"Generate a profile for: {customer.customer_name}\n\n"
                f"Use every labeled research section below (RAG, CRM, Web, LinkedIn).\n\n"
                f"{context}"
            ),
        }
    ]
    
    # Step 8: Get AI response (three-tier fallback in ai_service.ai_chat)
    try:
        profile_text = gemini_chat(messages, max_tokens=16384)
        if not profile_text or not profile_text.strip():
            raise RuntimeError("AI service returned empty response. Please check OPENAI_API_KEY configuration.")
    except Exception as e:
        error_msg = f"Failed to generate AI profile: {str(e)}"
        logging.error(error_msg)
        raise RuntimeError(error_msg)
    
    raw_profile_text = profile_text

    # Parse JSON matrix from raw response before plain-text sanitization removes it
    json_match = None
    json_patterns = [
        r'\{[^{}]*"strategic_fit_matrix"[^{}]*\{[^{}]*\}[^{}]*\}',  # Nested
        r'\{[^}]*"strategic_fit_matrix"[^}]*\}',  # Simple
    ]
    for pattern in json_patterns:
        json_match = re.search(pattern, raw_profile_text, re.IGNORECASE | re.DOTALL)
        if json_match:
            break
    
    # If no match, try to find any JSON block at the end of the response
    if not json_match:
        json_candidates = re.findall(
            r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}',
            raw_profile_text,
            re.IGNORECASE | re.DOTALL,
        )
        for candidate in reversed(json_candidates):
            if "strategic_fit_matrix" in candidate.lower():
                json_match = re.search(
                    re.escape(candidate), raw_profile_text, re.IGNORECASE | re.DOTALL
                )
                break

    profile_text = sanitize_profile_plain_text(raw_profile_text)

    if json_match:
        try:
            json_str = json_match.group(0)
            parsed = json.loads(json_str)
            matrix = parsed.get("strategic_fit_matrix", {})
            
            # Build product_scores dict using the actual category names as keys
            product_scores = {}
            for cat in categories_list:
                json_key = category_json_keys[cat]
                # Try both the JSON key and the original category name
                score = matrix.get(json_key) or matrix.get(cat.lower()) or matrix.get(cat)
                if score is None:
                    score = 0
                try:
                    product_scores[cat] = max(0, min(3, int(score)))
                except (ValueError, TypeError):
                    product_scores[cat] = 0
        except (json.JSONDecodeError, ValueError, KeyError):
            # If parsing fails, set defaults for all categories
            product_scores = {cat: 0 for cat in categories_list}
    else:
        # If no JSON found, set defaults for all categories
        product_scores = {cat: 0 for cat in categories_list}
    
    # Update the customer record with product alignment scores and latest ICP text
    update_payload: Dict[str, Any] = {}
    if product_scores:
        update_payload["product_alignment_scores"] = product_scores
    # Also store the full ICP profile text so ICP workspace can read it directly
    update_payload["latest_profile_text"] = profile_text
    update_payload["latest_profile_updated_at"] = datetime.utcnow().isoformat()
    research_meta["conversation_archive_count"] = archive_added
    research_meta["pipeline_archive_count"] = pipeline_added
    research_meta["chatgpt_export_count"] = chatgpt_added
    research_meta["interactions_table_count"] = table_total
    update_payload["latest_profile_research_meta"] = research_meta

    if update_payload:
        try:
            supabase.table("customers").update(update_payload).eq(
                "customer_id", customer.customer_id
            ).execute()
        except Exception as exc:
            logging.warning(
                "Profile update without research_meta (run docs/0002_profile_research_meta.sql): %s",
                exc,
            )
            update_payload.pop("latest_profile_research_meta", None)
            supabase.table("customers").update(update_payload).eq(
                "customer_id", customer.customer_id
            ).execute()

    # Store as an interaction so the history view shows it.
    interaction_payload = InteractionCreate(
        input_text=f"System: AI profile generated for {customer.customer_name}",
        ai_response=profile_text,
        tds_id=None,
    )
    _ = create_interaction(
        customer_id=str(customer.customer_id),
        interaction_in=interaction_payload,
        user_id=user_id,
    )

    # Log into the RAG `conversation` table as a combined entry.
    try:
        combined_text = (
            f"Customer: {customer.customer_name}\n"
            f"Display ID: {customer.display_id or '—'}\n"
            f"AI-generated CRM profile:\n{profile_text}"
        )
        embedding = gemini_embed(combined_text)
        metadata = {
            "customer_id": str(customer.customer_id),
            "customer_name": customer.customer_name,
            "source": "customer_profile",
            "tds_id": None,
            "user_id": user_id,
        }
        log_conversation_to_rag(
            combined_text,
            embedding=embedding,
            metadata=metadata,
        )
    except Exception:
        # Don't block profile building if RAG logging fails.
        pass

    # Return updated customer with scores
    updated_response = supabase.table("customers").select("*").eq("customer_id", customer.customer_id).execute()
    if updated_response.data:
        return Customer(**updated_response.data[0])
    return customer


# =============================
# INTERACTION SERVICES
# =============================


# Pagination when loading full CRM history for ICP / analysis
INTERACTIONS_PAGE_SIZE = 500
INTERACTIONS_MAX_FETCH = 2000


def _supabase_for_interaction_reads() -> Client:
    """
    Use service role for reads so Row Level Security does not hide historical rows.
    CRM UI and ICP analysis must see the full interactions table.
    """
    return get_supabase_service_client()


def _metadata_matches_customer(meta: Any, customer_id: str) -> bool:
    if not isinstance(meta, dict):
        return False
    cid = str(customer_id)
    for key in ("customer_id", "customerId", "customer_uuid"):
        val = meta.get(key)
        if val is not None and str(val) == cid:
            return True
    return False


def get_conversation_logs_for_customer(
    customer_id: str,
    *,
    max_rows: int = 2000,
) -> List[Dict[str, Any]]:
    """
    RAG archive (`conversation` table). Older chats may exist here without a row in
    `interactions` (e.g. legacy Streamlit imports, pipeline-only saves, or failed inserts).
    """
    supabase = _supabase_for_interaction_reads()
    cid = str(customer_id)
    collected: List[Dict[str, Any]] = []
    offset = 0
    page_size = 500

    def _append_row(row: Dict[str, Any]) -> None:
        meta = row.get("metadata") or {}
        if not isinstance(meta, dict):
            meta = {}
        collected.append(
            {
                "id": row.get("id"),
                "content": row.get("content") or "",
                "created_at": row.get("created_at"),
                "metadata": meta,
            }
        )

    # Prefer DB-side filter (fast, complete) instead of scanning the whole table in Python.
    while len(collected) < max_rows:
        try:
            resp = (
                supabase.table("conversation")
                .select("id, content, metadata, created_at")
                .eq("metadata->>customer_id", cid)
                .order("created_at", desc=True)
                .range(offset, offset + page_size - 1)
                .execute()
            )
        except Exception as exc:
            logging.warning(
                "conversation filter metadata->>customer_id failed for %s: %s",
                cid,
                exc,
            )
            break
        batch = resp.data or []
        if not batch:
            break
        for row in batch:
            _append_row(row)
            if len(collected) >= max_rows:
                break
        if len(batch) < page_size:
            break
        offset += page_size

    if collected:
        return collected

    # Fallback: paginate entire table and match flexible metadata keys (legacy rows).
    offset = 0
    while len(collected) < max_rows:
        resp = (
            supabase.table("conversation")
            .select("id, content, metadata, created_at")
            .order("created_at", desc=True)
            .range(offset, offset + page_size - 1)
            .execute()
        )
        batch = resp.data or []
        if not batch:
            break
        for row in batch:
            meta = row.get("metadata") or {}
            if not _metadata_matches_customer(meta, cid):
                continue
            _append_row(row)
            if len(collected) >= max_rows:
                break
        if len(batch) < page_size:
            break
        offset += page_size

    return collected


def get_pipeline_chat_history_for_customer(
    customer_id: str,
    *,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    max_rows: int = INTERACTIONS_MAX_FETCH,
) -> List[Interaction]:
    """
    Pipeline AI chats are stored in sales_pipeline.ai_interactions (JSON), not in
    public.interactions. They are also logged to conversation for RAG — merge includes both.
    """
    supabase = _supabase_for_interaction_reads()
    resp = (
        supabase.table("sales_pipeline")
        .select("id, customer_id, ai_interactions, created_at, updated_at")
        .eq("customer_id", customer_id)
        .execute()
    )
    rows = resp.data or []
    collected: List[Interaction] = []

    for pipeline in rows:
        pipeline_id = pipeline.get("id")
        raw = pipeline.get("ai_interactions")
        if isinstance(raw, str):
            try:
                entries = json.loads(raw)
            except json.JSONDecodeError:
                entries = []
        elif isinstance(raw, list):
            entries = raw
        else:
            entries = []

        for idx, entry in enumerate(entries):
            if not isinstance(entry, dict):
                continue
            created_at = (
                entry.get("timestamp")
                or entry.get("created_at")
                or pipeline.get("updated_at")
                or pipeline.get("created_at")
            )
            if not _row_in_date_range(created_at, start_date, end_date):
                continue
            input_text = (entry.get("user_input") or entry.get("input_text") or "").strip()
            ai_response = (entry.get("ai_response") or "").strip()
            if not input_text and not ai_response:
                continue
            synthetic_id = uuid5(
                NAMESPACE_URL,
                f"pipeline-{pipeline_id}-{idx}-{created_at}",
            )
            collected.append(
                Interaction(
                    id=synthetic_id,
                    customer_id=UUIDType(str(customer_id)),
                    input_text=input_text or None,
                    ai_response=ai_response or None,
                    created_at=created_at,
                    history_source="pipeline",
                )
            )
            if len(collected) >= max_rows:
                break
        if len(collected) >= max_rows:
            break

    collected.sort(key=lambda r: str(r.created_at or ""), reverse=True)
    return collected


def _parse_conversation_content(content: str) -> tuple[str, str]:
    """Extract user + AI parts from a RAG conversation row."""
    text = (content or "").strip()
    if not text:
        return "", ""

    patterns = [
        (r"(?:^|\n)Q:\s*(.+?)(?=\nA:|\Z)", r"(?:^|\n)A:\s*(.+?)(?=\nQ:|\nCustomer:|\Z)"),
        (r"(?:^|\n)Input:\s*(.+?)(?=\nOutput:|\Z)", r"(?:^|\n)Output:\s*(.+?)(?=\nInput:|\Z)"),
    ]
    for q_pat, a_pat in patterns:
        qm = re.search(q_pat, text, re.IGNORECASE | re.DOTALL)
        am = re.search(a_pat, text, re.IGNORECASE | re.DOTALL)
        if qm or am:
            return (
                (qm.group(1).strip() if qm else ""),
                (am.group(1).strip() if am else ""),
            )

    if text.lower().startswith("customer:"):
        text = re.sub(r"^Customer:\s*[^\n]+\n?", "", text, flags=re.IGNORECASE).strip()

    return ("[RAG conversation archive]", text)


def _interaction_fingerprint(input_text: str, ai_response: str) -> str:
    return f"{(input_text or '').strip()[:180]}|{(ai_response or '').strip()[:180]}"


def _row_in_date_range(
    created_at: Any,
    start_date: Optional[str],
    end_date: Optional[str],
) -> bool:
    if not start_date and not end_date:
        return True
    if not created_at:
        return False
    raw = str(created_at)
    day = raw[:10]
    if start_date and day < start_date:
        return False
    if end_date and day > end_date:
        return False
    return True


def merge_customer_interaction_history(
    customer_id: str,
    *,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    max_rows: int = INTERACTIONS_MAX_FETCH,
) -> tuple[List[Interaction], int, int, int, int]:
    """
    Unified timeline: interactions + conversation + pipeline JSON + ChatGPT exports (deduped).
    Returns (merged, table_count, conversation-only, pipeline-only, chatgpt_export-only).
    """
    customer_row = get_customer_by_id(customer_id)
    customer_name = customer_row.customer_name if customer_row else ""

    table_rows = get_all_interactions_for_customer(
        customer_id,
        max_rows=max_rows,
        start_date=start_date,
        end_date=end_date,
    )
    fingerprints = {
        _interaction_fingerprint(it.input_text or "", it.ai_response or "")
        for it in table_rows
    }

    merged: List[Interaction] = []
    for it in table_rows:
        payload = it.model_dump()
        payload["history_source"] = "interactions"
        merged.append(Interaction(**payload))

    archive_added = 0
    for row in get_conversation_logs_for_customer(customer_id, max_rows=max_rows):
        if not _row_in_date_range(row.get("created_at"), start_date, end_date):
            continue
        input_text, ai_response = _parse_conversation_content(row.get("content") or "")
        fp = _interaction_fingerprint(input_text, ai_response)
        if fp in fingerprints:
            continue
        fingerprints.add(fp)

        conv_id = row.get("id")
        try:
            synthetic_id = UUIDType(str(conv_id)) if conv_id else uuid5(NAMESPACE_URL, "conversation-empty")
        except (ValueError, TypeError):
            synthetic_id = uuid5(NAMESPACE_URL, f"conversation-{conv_id}")

        merged.append(
            Interaction(
                id=synthetic_id,
                customer_id=UUIDType(str(customer_id)),
                input_text=input_text or None,
                ai_response=ai_response or None,
                created_at=row.get("created_at"),
                history_source="conversation",
            )
        )
        archive_added += 1

    pipeline_added = 0
    for it in get_pipeline_chat_history_for_customer(
        customer_id,
        start_date=start_date,
        end_date=end_date,
        max_rows=max_rows,
    ):
        fp = _interaction_fingerprint(it.input_text or "", it.ai_response or "")
        if fp in fingerprints:
            continue
        fingerprints.add(fp)
        merged.append(it)
        pipeline_added += 1

    chatgpt_added = 0
    if customer_name:
        for row in get_chatgpt_export_archives_for_customer(
            customer_id, customer_name, max_rows=30
        ):
            if not _row_in_date_range(row.get("created_at"), start_date, end_date):
                continue
            input_text, ai_response, parsed_time = _parse_chatgpt_export_row(
                row.get("content") or ""
            )
            fp = _interaction_fingerprint(input_text, ai_response)
            if fp in fingerprints:
                continue
            fingerprints.add(fp)

            conv_id = row.get("id")
            try:
                synthetic_id = (
                    UUIDType(str(conv_id))
                    if conv_id
                    else uuid5(NAMESPACE_URL, "chatgpt-export-empty")
                )
            except (ValueError, TypeError):
                synthetic_id = uuid5(NAMESPACE_URL, f"chatgpt-export-{conv_id}")

            merged.append(
                Interaction(
                    id=synthetic_id,
                    customer_id=UUIDType(str(customer_id)),
                    input_text=input_text or None,
                    ai_response=ai_response or None,
                    created_at=parsed_time or row.get("created_at"),
                    history_source="chatgpt_export",
                )
            )
            chatgpt_added += 1

    merged.sort(
        key=lambda r: str(r.created_at or ""),
        reverse=True,
    )
    if len(merged) > max_rows:
        merged = merged[:max_rows]
    return merged, len(table_rows), archive_added, pipeline_added, chatgpt_added


def audit_customer_interaction_sources(customer_id: str) -> Dict[str, Any]:
    """Counts by month in interactions vs conversation vs pipeline JSON for diagnostics."""
    interactions = get_all_interactions_for_customer(customer_id)
    conversations = get_conversation_logs_for_customer(customer_id)
    pipeline_rows = get_pipeline_chat_history_for_customer(customer_id)

    def bucket(rows: List[Any], date_fn) -> Dict[str, int]:
        counts: Dict[str, int] = defaultdict(int)
        for row in rows:
            raw = date_fn(row)
            key = str(raw)[:7] if raw else "unknown"
            counts[key] += 1
        return dict(sorted(counts.items()))

    def count_may(rows: List[Any], date_fn) -> int:
        return sum(
            1
            for r in rows
            if date_fn(r) and str(date_fn(r))[5:7] == "05"
        )

    merged, table_n, conv_n, pipe_n, gpt_n = merge_customer_interaction_history(customer_id)

    return {
        "customer_id": customer_id,
        "interactions_table": "public.interactions",
        "conversation_table": "public.conversation",
        "pipeline_table": "public.sales_pipeline.ai_interactions",
        "interactions_total": len(interactions),
        "conversation_total": len(conversations),
        "pipeline_total": len(pipeline_rows),
        "merged_total": len(merged),
        "interactions_by_month": bucket(
            interactions, lambda r: getattr(r, "created_at", None)
        ),
        "conversation_by_month": bucket(
            conversations, lambda r: r.get("created_at")
        ),
        "pipeline_by_month": bucket(
            pipeline_rows, lambda r: getattr(r, "created_at", None)
        ),
        "may_interactions": count_may(interactions, lambda r: getattr(r, "created_at", None)),
        "may_conversation": count_may(conversations, lambda r: r.get("created_at")),
        "may_pipeline": count_may(pipeline_rows, lambda r: getattr(r, "created_at", None)),
        "may_merged": count_may(merged, lambda r: getattr(r, "created_at", None)),
        "merged_table_rows": table_n,
        "merged_conversation_only": conv_n,
        "merged_pipeline_only": pipe_n,
        "chatgpt_export_total": gpt_n,
        "merged_chatgpt_export_only": gpt_n,
    }


def get_all_interactions_for_customer(
    customer_id: str,
    *,
    max_rows: int = INTERACTIONS_MAX_FETCH,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> List[Interaction]:
    """Load every interaction for a customer (paginated Supabase reads)."""
    collected: List[Interaction] = []
    offset = 0
    while len(collected) < max_rows:
        page_size = min(INTERACTIONS_PAGE_SIZE, max_rows - len(collected))
        page = get_interactions_for_customer(
            customer_id,
            limit=page_size,
            offset=offset,
            start_date=start_date,
            end_date=end_date,
        )
        if not page:
            break
        collected.extend(page)
        if len(page) < page_size:
            break
        offset += len(page)
    return collected


def get_interactions_for_customer(
    customer_id: str,
    limit: int = 100,
    offset: int = 0,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> List[Interaction]:
    """Get interactions for a specific customer, newest first.
    
    Args:
        customer_id: Customer UUID
        limit: Maximum number of interactions to return
        offset: Number of interactions to skip
        start_date: Optional ISO date string (YYYY-MM-DD) - filter interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter interactions up to this date
    """
    supabase = _supabase_for_interaction_reads()

    query = (
        supabase.table("interactions")
        .select("*")
        .eq("customer_id", customer_id)
    )

    # Apply date filters if provided
    if start_date:
        query = query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        query = query.lte("created_at", f"{end_date}T23:59:59")

    response = (
        query.order("created_at", desc=True)
        .limit(limit)
        .offset(offset)
        .execute()
    )

    return [Interaction(**row) for row in (response.data or [])]


def get_interactions_count_for_customer(
    customer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> int:
    """Get total interaction count for a customer (for pagination).
    
    Args:
        customer_id: Customer UUID
        start_date: Optional ISO date string (YYYY-MM-DD) - filter interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter interactions up to this date
    """
    supabase = _supabase_for_interaction_reads()

    query = (
        supabase.table("interactions")
        .select("id", count="exact")
        .eq("customer_id", customer_id)
    )

    # Apply date filters if provided
    if start_date:
        query = query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        query = query.lte("created_at", f"{end_date}T23:59:59")

    response = query.execute()

    return response.count if getattr(response, "count", None) is not None else 0


def get_interaction_by_id(interaction_id: str) -> Optional[Interaction]:
    """Fetch a single interaction by its UUID."""
    supabase = _supabase_for_interaction_reads()

    response = (
        supabase.table("interactions")
        .select("*")
        .eq("id", interaction_id)
        .execute()
    )

    if response.data:
        return Interaction(**response.data[0])
    return None


def create_interaction(
    customer_id: str, interaction_in: InteractionCreate, user_id: Optional[str] = None
) -> Interaction:
    """Create a new interaction linked to a customer (and optional user)."""
    supabase: Client = get_supabase_client()

    payload = interaction_in.model_dump(exclude_unset=True)
    payload["customer_id"] = customer_id
    if user_id:
        payload["user_id"] = user_id

    response = supabase.table("interactions").insert(payload).execute()

    if not response.data:
        raise RuntimeError("Failed to create interaction")
    interaction_row = response.data[0]

    # ------------------------------------------------------
    # Enqueue async ICP profile update job (deduplicated)
    # ------------------------------------------------------
    try:
        # Basic dedupe: if there is already a queued/processing job
        # for this customer, don't enqueue another one.
        existing_jobs = (
            supabase.table("profile_update_jobs")
            .select("id, status")
            .eq("customer_id", customer_id)
            .in_("status", ["queued", "processing"])
            .limit(1)
            .execute()
        )

        if not existing_jobs.data:
            job_payload: Dict[str, Any] = {
                "customer_id": customer_id,
                "interaction_id": interaction_row.get("id"),
                "status": "queued",
            }

            # If this interaction was linked to a specific pipeline, propagate it
            pipeline_id = interaction_row.get("pipeline_id")
            if pipeline_id:
                job_payload["pipeline_id"] = pipeline_id

            supabase.table("profile_update_jobs").insert(job_payload).execute()
    except Exception as e:
        # Fail-safe: interaction creation should NOT fail just because
        # job enqueue failed. Log and continue.
        logging.warning(
            f"Failed to enqueue profile_update_jobs for customer {customer_id}: {e}"
        )

    try:
        customer = get_customer_by_id(customer_id)
        if customer and not (interaction_in.input_text or "").strip().startswith(
            "[telegram_backfill]"
        ):
            notify_interaction_saved(
                customer_name=customer.customer_name,
                customer_id=str(customer_id),
                input_text=interaction_in.input_text or "",
                ai_response=interaction_in.ai_response or "",
                created_at=str(interaction_row.get("created_at") or ""),
                source="crm",
            )
    except Exception as e:
        logging.warning("Telegram notify skipped for %s: %s", customer_id, e)

    return Interaction(**interaction_row)


def update_customer_profile_text(
    customer_id: str, profile_update: CustomerProfileUpdate
) -> Customer:
    """Update the latest ICP profile text for a customer."""
    supabase: Client = get_supabase_client()

    # Ensure customer exists
    customer = get_customer_by_id(customer_id)
    if not customer:
        raise ValueError("Customer not found")

    now_iso = datetime.utcnow().isoformat()
    response = (
        supabase.table("customers")
        .update(
            {
                "latest_profile_text": profile_update.profile_text,
                "latest_profile_updated_at": now_iso,
            }
        )
        .eq("customer_id", customer_id)
        .execute()
    )

    if not response.data:
        raise RuntimeError("Failed to update customer profile text")

    return Customer(**response.data[0])


def add_customer_profile_feedback(
    customer_id: str, feedback_in: CustomerProfileFeedbackCreate
) -> CustomerProfileFeedback:
    """Store a rating/comment for a customer's ICP profile.
    
    If the feedback table doesn't exist yet, logs a warning and returns a mock response
    (graceful degradation - feedback is optional).
    """
    supabase: Client = get_supabase_client()

    # Basic validation: rating 1-5
    if feedback_in.rating < 1 or feedback_in.rating > 5:
        raise ValueError("rating must be between 1 and 5")

    # Ensure customer exists
    customer = get_customer_by_id(customer_id)
    if not customer:
        raise ValueError("Customer not found")

    try:
        payload = {
            "customer_id": customer_id,
            "rating": feedback_in.rating,
            "comment": feedback_in.comment,
        }

        response = supabase.table("customer_profile_feedback").insert(payload).execute()
        if not response.data:
            raise RuntimeError("Failed to insert feedback")

        row = response.data[0]
        return CustomerProfileFeedback(
            id=row["id"],
            customer_id=row["customer_id"],
            rating=row["rating"],
            comment=row.get("comment"),
            user_id=row.get("user_id"),
            created_at=row.get("created_at"),
        )
    except Exception as e:
        # If table doesn't exist (PGRST205), log and return a mock response
        # This allows the UI to work without the feedback table
        error_msg = str(e)
        if "PGRST205" in error_msg or "customer_profile_feedback" in error_msg.lower():
            logging.warning(
                f"Feedback table not found, feedback not saved. "
                f"To enable feedback, run: backend/scripts/add_customer_profile_feedback.sql. "
                f"Error: {error_msg}"
            )
            # Return a mock response so the frontend doesn't break
            from uuid import uuid4
            from datetime import datetime
            return CustomerProfileFeedback(
                id=str(uuid4()),
                customer_id=customer_id,
                rating=feedback_in.rating,
                comment=feedback_in.comment,
                user_id=None,
                created_at=datetime.utcnow().isoformat(),
            )
        # Re-raise other unexpected errors
        raise


def list_customer_profile_feedback(
    customer_id: str, limit: int = 10
) -> List[CustomerProfileFeedback]:
    """Return recent feedback entries for a customer's profile.
    
    If the feedback table doesn't exist yet, returns an empty list (graceful degradation).
    """
    supabase: Client = get_supabase_client()

    try:
        response = (
            supabase.table("customer_profile_feedback")
            .select("*")
            .eq("customer_id", customer_id)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
        )

        rows = response.data or []
        return [
            CustomerProfileFeedback(
                id=row["id"],
                customer_id=row["customer_id"],
                rating=row["rating"],
                comment=row.get("comment"),
                user_id=row.get("user_id"),
                created_at=row.get("created_at"),
            )
            for row in rows
        ]
    except Exception as e:
        # If table doesn't exist (PGRST205) or any other error, return empty list
        # This allows the feature to work without the feedback table
        error_msg = str(e)
        if "PGRST205" in error_msg or "customer_profile_feedback" in error_msg.lower():
            logging.warning(f"Feedback table not found, returning empty list. Error: {error_msg}")
            return []
        # Re-raise other unexpected errors
        raise


def update_interaction(
    interaction_id: str, interaction_in: InteractionUpdate
) -> Optional[Interaction]:
    """Update an existing interaction and return the updated record."""
    supabase: Client = get_supabase_client()

    update_data = interaction_in.model_dump(exclude_unset=True)
    if not update_data:
        # Nothing to update; just return the current record
        return get_interaction_by_id(interaction_id)

    response = (
        supabase.table("interactions")
        .update(update_data)
        .eq("id", interaction_id)
        .execute()
    )

    if response.data:
        return Interaction(**response.data[0])
    return None


def delete_interaction(interaction_id: str) -> None:
    """Delete an interaction. Raises on failure."""
    supabase: Client = get_supabase_client()

    supabase.table("interactions").delete().eq("id", interaction_id).execute()


def auto_fill_sales_stage_for_customer(customer_id: str) -> Optional[str]:
    """
    Analyze and set sales stage for a single customer based on their interaction history.
    
    Args:
        customer_id: The customer ID to analyze
        
    Returns:
        The new sales stage (1-7) if successful, None if customer not found or has no interactions
    """
    customer = get_customer_by_id(customer_id)
    if not customer:
        return None
    
    # If customer already has a sales stage, don't overwrite (user can manually edit)
    if customer.sales_stage:
        return customer.sales_stage
    
    try:
        # Get all interactions for this customer
        interactions = get_all_interactions_for_customer(customer_id, max_rows=200)
        
        if not interactions:
            # No interactions = Stage 1 (Prospecting)
            supabase: Client = get_supabase_client()
            supabase.table("customers").update({"sales_stage": "1"}).eq("customer_id", customer_id).execute()
            return "1"
        
        # Build context from interactions
        history_lines = []
        for it in interactions:
            user_part = (it.input_text or "").strip()
            ai_part = (it.ai_response or "").strip()
            if user_part or ai_part:
                history_lines.append(f"Q: {user_part}\nA: {ai_part}")
        
        past_context = "\n\n".join(history_lines) if history_lines else "No past interactions"
        
        # Use the most recent interaction for analysis
        latest = interactions[0]  # Already sorted newest first
        new_interaction = f"Q: {latest.input_text or ''}\nA: {latest.ai_response or ''}"
        
        # Analyze stage
        new_stage = analyze_sales_stage(new_interaction, past_context, current_stage=None)
        
        # Update customer
        supabase: Client = get_supabase_client()
        supabase.table("customers").update({"sales_stage": new_stage}).eq("customer_id", customer_id).execute()
        
        return new_stage
        
    except Exception as e:
        print(f"Error auto-filling sales stage for customer {customer_id}: {e}")
        return None


def backfill_sales_stages_for_all_customers() -> Dict[str, Any]:
    """
    Analyze and set sales stages for all customers that don't have one yet.
    Uses their interaction history to determine the current stage.
    
    Returns:
        Dict with counts: {"updated": N, "skipped": M, "errors": K}
    """
    supabase: Client = get_supabase_client()
    
    # Get all customers - we'll filter for null/empty sales_stage in Python
    response = (
        supabase.table("customers")
        .select("customer_id, customer_name, sales_stage")
        .execute()
    )
    
    all_customers = response.data or []
    # Filter for customers without sales_stage
    customers_to_update = [
        c for c in all_customers 
        if not c.get("sales_stage") or c.get("sales_stage") == ""
    ]
    
    results = {"updated": 0, "skipped": 0, "errors": 0}
    
    for customer_row in customers_to_update:
        customer_id = customer_row["customer_id"]
        customer_name = customer_row["customer_name"]
        
        try:
            # Get all interactions for this customer
            interactions = get_all_interactions_for_customer(customer_id, max_rows=200)
            
            if not interactions:
                # No interactions = Stage 1 (Prospecting)
                supabase.table("customers").update({"sales_stage": "1"}).eq("customer_id", customer_id).execute()
                results["updated"] += 1
                continue
            
            # Build context from interactions
            history_lines = []
            for it in interactions:
                user_part = (it.input_text or "").strip()
                ai_part = (it.ai_response or "").strip()
                if user_part or ai_part:
                    history_lines.append(f"Q: {user_part}\nA: {ai_part}")
            
            past_context = "\n\n".join(history_lines) if history_lines else "No past interactions"
            
            # Use the most recent interaction for analysis
            latest = interactions[0]  # Already sorted newest first
            new_interaction = f"Q: {latest.input_text or ''}\nA: {latest.ai_response or ''}"
            
            # Analyze stage
            new_stage = analyze_sales_stage(new_interaction, past_context, current_stage=None)
            
            # Update customer
            supabase.table("customers").update({"sales_stage": new_stage}).eq("customer_id", customer_id).execute()
            results["updated"] += 1
            
        except Exception as e:
            print(f"Error backfilling sales stage for customer {customer_id}: {e}")
            results["errors"] += 1
    
    return results


def _parse_interaction_datetime(value: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        normalized = value.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except (TypeError, ValueError):
        return None


def _week_start_label(dt: datetime) -> str:
    monday = dt - timedelta(days=dt.weekday())
    return monday.strftime("%Y-%m-%d")


def get_dashboard_metrics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> DashboardMetrics:
    """
    Get dashboard metrics including customer counts, interaction counts, and sales stage distribution.
    
    Args:
        start_date: Optional ISO date string (YYYY-MM-DD) - filter interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter interactions up to this date
    
    Returns:
        DashboardMetrics with all calculated metrics
    """
    supabase: Client = get_supabase_client()
    
    # All customers (for quiet list and totals)
    customers_response = supabase.table("customers").select(
        "customer_id, customer_name, display_id, sales_stage"
    ).execute()
    customer_rows = customers_response.data or []
    total_customers = len(customer_rows)

    # Get interactions count with date filtering
    interactions_query = supabase.table("interactions").select("id", count="exact")
    if start_date:
        interactions_query = interactions_query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        interactions_query = interactions_query.lte("created_at", f"{end_date}T23:59:59")
    
    interactions_response = interactions_query.execute()
    total_interactions = interactions_response.count if getattr(interactions_response, "count", None) is not None else 0
    
    # Get distinct customers with interactions (within date range if specified)
    customers_with_interactions_query = supabase.table("interactions").select("customer_id")
    if start_date:
        customers_with_interactions_query = customers_with_interactions_query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        customers_with_interactions_query = customers_with_interactions_query.lte("created_at", f"{end_date}T23:59:59")
    
    customers_with_interactions_response = customers_with_interactions_query.execute()
    unique_customer_ids = set(
        row.get("customer_id")
        for row in (customers_with_interactions_response.data or [])
        if row.get("customer_id")
    )
    customers_with_interactions = len(unique_customer_ids)

    quiet_customers: List[QuietCustomerSummary] = []
    for row in customer_rows:
        cid = row.get("customer_id")
        if cid and cid not in unique_customer_ids:
            quiet_customers.append(
                QuietCustomerSummary(
                    customer_id=cid,
                    customer_name=(row.get("customer_name") or "Unknown").strip(),
                    display_id=row.get("display_id"),
                )
            )
    quiet_customers.sort(key=lambda c: c.customer_name.lower())

    # Weekly interaction counts (uses same date filters as metrics)
    interactions_weekly_query = supabase.table("interactions").select("created_at")
    if start_date:
        interactions_weekly_query = interactions_weekly_query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        interactions_weekly_query = interactions_weekly_query.lte("created_at", f"{end_date}T23:59:59")
    interactions_weekly_response = interactions_weekly_query.execute()
    week_counts: Dict[str, int] = defaultdict(int)
    for row in interactions_weekly_response.data or []:
        dt = _parse_interaction_datetime(row.get("created_at") or "")
        if dt:
            week_counts[_week_start_label(dt)] += 1
    interactions_by_week = [
        WeeklyInteractionCount(week_start=week, count=count)
        for week, count in sorted(week_counts.items())
    ]

    sales_stages_distribution: Dict[str, int] = {
        "1": 0, "2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0
    }
    for row in customer_rows:
        stage = row.get("sales_stage")
        if stage and stage in sales_stages_distribution:
            sales_stages_distribution[stage] = sales_stages_distribution.get(stage, 0) + 1

    return DashboardMetrics(
        total_customers=total_customers,
        total_interactions=total_interactions,
        customers_with_interactions=customers_with_interactions,
        sales_stages_distribution=sales_stages_distribution,
        quiet_customers=quiet_customers,
        interactions_by_week=interactions_by_week,
    )


def analyze_sales_stage(
    new_interaction: str,
    past_context: str,
    current_stage: Optional[str] = None,
) -> str:
    """
    Analyze the sales stage based on Brian Tracy's 7-stage process.
    
    Returns the current stage number (1-7) as a string.
    """
    system_prompt = f"""You are "LeanChem 7-Stage Sales Tracker" based on Brian Tracy's sales process.

Analyze the customer interaction and determine which of the 7 stages they are currently in:

1. Prospecting - Customer identified, initial contact made
2. Rapport - Trust built, relationship established
3. Needs Analysis - Customer shares requirements, pain points identified
4. Presenting Solution - Product/service proposal presented
5. Handling Objections - Addressing concerns, negotiating terms
6. Closing - Finalizing deal, contract signed
7. Follow-up & Cross-sell - Post-sale support, upselling opportunities

CURRENT STAGE (if known): {current_stage or "None"}

PAST CONTEXT:
\"\"\"{past_context}\"\"\"

NEW INTERACTION:
\"\"\"{new_interaction}\"\"\"

TASK:
- Analyze the NEW INTERACTION in context of PAST CONTEXT
- Determine the CURRENT stage (1-7) based on evidence
- If CURRENT STAGE is provided, only advance if there's strong new evidence
- Return ONLY the stage number (1, 2, 3, 4, 5, 6, or 7) - nothing else

OUTPUT FORMAT:
Return only a single digit: 1, 2, 3, 4, 5, 6, or 7
"""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": "What is the current sales stage? Return only the number (1-7)."},
    ]

    try:
        response = gemini_chat(messages).strip()
        # Extract just the number if there's extra text
        import re
        match = re.search(r'\b([1-7])\b', response)
        if match:
            stage_num = match.group(1)
            # Validate it's 1-7
            if stage_num in SALES_STAGES:
                return stage_num
        # Default to stage 1 if parsing fails
        return "1"
    except Exception:
        # Default to stage 1 on error
        return "1"


def chat_with_customer(
    customer_id: str,
    input_text: str,
    tds_id: Optional[str] = None,
    user_id: Optional[str] = None,
    file_url: Optional[str] = None,
    file_type: Optional[str] = None,
    file_content: Optional[str] = None,
) -> Interaction:
    """
    Run an AI chat turn for a specific customer:
    - Calls Gemini to generate a response
    - Stores the turn in `interactions` table
    - Logs a combined Q/A entry into `conversation` (RAG) with embedding
    """
    supabase: Client = get_supabase_client()

    # 1) Ensure customer exists
    customer = get_customer_by_id(customer_id)
    if not customer:
        raise RuntimeError("Customer not found")

    # 2) Fetch recent interactions to give the AI richer CRM context
    recent_interactions, _, _, _, _ = merge_customer_interaction_history(
        customer_id, max_rows=100
    )
    # Oldest first in the prompt so the story reads naturally
    recent_interactions = list(reversed(recent_interactions))

    history_lines: list[str] = []
    for i, it in enumerate(recent_interactions, start=1):
        user_part = (it.input_text or "").strip()
        ai_part = (it.ai_response or "").strip()
        if not user_part and not ai_part:
            continue
        history_lines.append(
            f"Interaction {i}:\n"
            f"Input: {user_part or '[no input logged]'}\n"
            f"Output: {ai_part or '[no AI response logged]'}"
        )

    memories_str = "\n\n".join(history_lines) if history_lines else "No past interactions yet."

    # 2.5) Fetch sales pipelines for this customer to provide pipeline context
    sales_pipelines = []
    try:
        # Import here to avoid circular import
        from app.services.sales_pipeline_service import list_sales_pipelines
        sales_pipelines = list_sales_pipelines(
            limit=20,
            customer_id=str(customer.customer_id),
        )
    except Exception:
        pass  # Don't fail if pipeline service is unavailable
    
    pipeline_context = ""
    if sales_pipelines:
        pipeline_context = f"\n\nSales Pipeline Information ({len(sales_pipelines)} pipeline(s)):\n"
        for idx, p in enumerate(sales_pipelines[:10], 1):  # Show up to 10 pipelines
            amount_str = f"{p.amount or 0:,.2f} {p.currency or 'USD'}" if p.amount else "Not set"
            pipeline_context += f"- Pipeline {idx}: Stage: {p.stage}, Amount: {amount_str}, Expected Close: {p.expected_close_date or 'Not set'}\n"
    else:
        pipeline_context = "\n\nSales Pipeline Information: No active pipelines found for this customer.\n"

    customer_context = (
        f"Customer name: {customer.customer_name}\n"
        f"Display ID: {customer.display_id or '—'}\n"
        f"Customer ID: {customer.customer_id}\n"
        f"Total recorded interactions: {len(recent_interactions)}"
    )

    # 3) Prepare messages for Gemini using the same style as the Streamlit CRM chat
    system_prompt = f"""
You are a helpful AI assistant specialized in chemical trading and CRM.
If the user asks about a specific customer, use the customer's most relevant past interactions below.
Also use the provided memories, sales pipeline information, and any relevant conversations from the database.
If you don't find relevant information, say so and reason transparently.

Customer context:
{customer_context}
{pipeline_context}

User Memories:
{memories_str}
"""

    # 3.5) Add file content to context if provided
    user_content = input_text
    if file_content:
        user_content = f"{input_text}\n\n--- Attached File Content ---\n{file_content}"

    messages = [
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": user_content,
        },
    ]

    # 4) Call Gemini
    ai_response = gemini_chat(messages)

    # 5) Store in interactions table
    interaction_payload = InteractionCreate(
        input_text=input_text,
        ai_response=ai_response,
        tds_id=tds_id,
        file_url=file_url,
        file_type=file_type,
    )
    interaction = create_interaction(customer_id, interaction_payload, user_id=user_id)

    # 6) Log to RAG `conversation` table with embedding
    try:
        combined_text = f"Customer: {customer.customer_name}\nQ: {input_text}\nA: {ai_response}"
        embedding = gemini_embed(combined_text)
        metadata = {
            "customer_id": str(customer.customer_id),
            "customer_name": customer.customer_name,
            "source": "crm_chat",
            "tds_id": tds_id,
            "user_id": user_id,
        }
        log_conversation_to_rag(combined_text, embedding=embedding, metadata=metadata)
    except Exception:
        # Don't block the main interaction flow if RAG logging fails
        pass

    # 7) Analyze and update sales stage
    try:
        # Build context for stage analysis
        stage_context = f"{customer_context}\n\nRecent Interactions:\n{memories_str}"
        new_stage = analyze_sales_stage(
            new_interaction=f"Q: {input_text}\nA: {ai_response}",
            past_context=stage_context,
            current_stage=customer.sales_stage,
        )
        
        # Update customer's sales stage if it changed
        if new_stage != customer.sales_stage:
            supabase.table("customers").update({"sales_stage": new_stage}).eq("customer_id", customer_id).execute()
    except Exception:
        # Don't block the main interaction flow if stage analysis fails
        pass

    return interaction


def generate_quote_excel(body: QuoteDraftRequest) -> str:
    """
    Generate an AI-enhanced Excel quotation file based on a template.

    Baracoda format:
      - B12 = customer name
      - C20, C21, ... = product names (starting at row 20)
      - I20, I21, ... = quantities
      - J20, J21, ... = unit prices
      - L20, L21, ... = formulas =J20*I20, =J21*I21, etc.
      - L30 = SUM(L20:L{last_row})*0.15 (VAT)
      - L31 = SUM(L20:L30) (total incl. VAT)

    Betchem format:
      - A4 = company name
      - E4 = date
      - B8, B9, ... = product names (starting at row 8)
      - C8, C9, ... = unit measurement (kg, mt, etc.)
      - D8, D9, ... = quantities
      - E8, E9, ... = unit prices before VAT
      - F8, F9, ... = formulas =D8*E8, =D9*E9, etc.
      - F12 = SUM(F8:F{last_row}) (subtotal)
      - F13 = F12*0.15 (VAT)
      - F14 = F12+F13 (total incl. VAT)
    """
    # 1) Locate template on disk
    fmt = body.format.lower()
    if fmt == "baracoda":
        filename = "Baracoda.xlsx"
    elif fmt == "betchem":
        filename = "Betchem.xlsx"
    else:
        raise RuntimeError(f"Unsupported quote format: {body.format}")

    template_path = Path(__file__).resolve().parents[3] / "qoute_format" / filename
    if not template_path.exists():
        raise RuntimeError(f"Template file not found: {template_path}")

    # 2) Open template and get the first/active sheet
    wb = load_workbook(template_path)
    ws = wb.active

    # 3) Parse quantities and prices for all products
    parsed_products = []
    for p in body.products:
        # Parse quantity
        try:
            qty_val = float(p.quantity)
        except Exception:
            qty_val = None

        # Parse unit price
        unit_price_val: Optional[float] = None
        if p.target_price is not None:
            try:
                unit_price_val = float(p.target_price)
            except Exception:
                # Try to extract leading numeric part like "1200" from "1200 USD/MT"
                m = re.match(r"\s*([0-9]+(?:\.[0-9]+)?)", str(p.target_price))
                if m:
                    try:
                        unit_price_val = float(m.group(1))
                    except Exception:
                        unit_price_val = None

        parsed_products.append({
            "name": p.chemical_type_name,
            "unit": p.unit,
            "quantity": qty_val if qty_val is not None else p.quantity,
            "unit_price": unit_price_val if unit_price_val is not None else (p.target_price or ""),
        })

    # 4) Apply format-specific mappings
    if fmt == "baracoda":
        # Baracoda format
        ws["B12"] = body.customer_name

        start_row = 20
        last_product_row = start_row + len(parsed_products) - 1

        for idx, p in enumerate(parsed_products):
            row = start_row + idx
            ws[f"C{row}"] = p["name"]
            ws[f"I{row}"] = p["quantity"]
            ws[f"J{row}"] = p["unit_price"]
            ws[f"L{row}"] = f"=J{row}*I{row}"

        ws["L30"] = f"=SUM(L{start_row}:L{last_product_row})*0.15"
        ws["L31"] = f"=SUM(L{start_row}:L30)"

        # Write terms and conditions to B34 for Baracoda
        # Use provided terms or default
        terms_text = body.terms_and_conditions or "Terms and conditions:\nMinium Order Quantity: 1000 KG Per Product\nPayment: Advance Payment is 50% 30% when goods are delivered at Moyale and & Balance Payment is 20% on Delivery."
        cell = ws["B34"]
        cell.value = terms_text
        # Enable text wrapping for multi-line content
        from openpyxl.styles import Alignment
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    elif fmt == "betchem":
        # Betchem format
        ws["A4"] = body.customer_name
        # Write current date to E4
        from datetime import datetime
        ws["E4"] = datetime.now().strftime("%Y-%m-%d")

        start_row = 8
        last_product_row = start_row + len(parsed_products) - 1

        for idx, p in enumerate(parsed_products):
            row = start_row + idx
            ws[f"B{row}"] = p["name"]
            ws[f"C{row}"] = p["unit"]  # Unit measurement (kg, mt, etc.)
            ws[f"D{row}"] = p["quantity"]
            ws[f"E{row}"] = p["unit_price"]
            ws[f"F{row}"] = f"=D{row}*E{row}"

        # F12 = sum of all product subtotals
        ws["F12"] = f"=SUM(F{start_row}:F{last_product_row})"
        # F13 = VAT (15%)
        ws["F13"] = "=F12*0.15"
        # F14 = total including VAT
        ws["F14"] = "=F12+F13"

        # Write terms and conditions to A16 for Betchem
        # Use provided terms or default
        terms_text = body.terms_and_conditions or "Terms and conditions:\n- For items currently in stock, the advance payment is 100 %"
        cell = ws["A16"]
        cell.value = terms_text
        # Enable text wrapping for multi-line content
        from openpyxl.styles import Alignment
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 5) Save to a temporary file and return its path
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"_{filename}")
    tmp_path = Path(tmp.name)
    tmp.close()
    wb.save(tmp_path)

    return str(tmp_path)